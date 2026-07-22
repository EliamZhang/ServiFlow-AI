# -*- coding: utf-8 -*-
"""Compare our classification output against illion labels.

Usage::

    python compare_labels.py [--output output/classification_report.xlsx]

The illion labels are ~95 % accurate but *not* ground truth —
disagreement does not always mean we are wrong.

Output (xlsx)
-------------
1. **Coverage** — who classified more rows?
2. **Confusion** — our finv_category × illion category pivot
3. **Counterparty** — coverage + exact-match rate
4. **Gaps** — rows only we / only illion classified
"""

from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent
SAMPLE_CSV = PROJECT_ROOT / "sample.csv"
DEFAULT_OUTPUT = PROJECT_ROOT / "output" / "classification_report.xlsx"
DEFAULT_RESULT = PROJECT_ROOT / "output" / "compare_result.xlsx"

# ── category name mapping ──────────────────────────────────────────────
#
# Two mapping tables for human review:
#   OUR_TO_ILLION  — our finv_category → illion category (drives agreement check)
#   ILLION_TO_OUR  — illion category → our finv_category(s)  (reverse reference)
#
# Exact name matches work automatically and are omitted from OUR_TO_ILLION.
# Categories in UNMAPPABLE_OUR_CATS have no illion equivalent and are
# excluded from agreement-rate calculation.

OUR_TO_ILLION: dict[str, str] = {
    # -- singular / plural --
    "fee":                   "Fees",
    "Credit Card Repayment": "Credit Card Repayments",

    # -- case --
    "centrelink":            "Centrelink",

    # -- income: our fine-grained → illion "Wages" --
    "salary_payg":           "Wages",
    "salary_packaging":      "Wages",
    "self_employed_gig":     "Wages",

    # -- liability: our fine-grained → illion coarse-grained --
    "BNPL":                  "Non SACC Loans",
    "Wage Advance":          "Non SACC Loans",
    "Personal Loan Unknown": "Non SACC Loans",
    "Contract Loans":        "Non SACC Loans",
    "LOC":                   "Non SACC Loans",
    "Home Loan":             "Non SACC Loans",
    "Car Loan":              "Non SACC Loans",
}

# Our categories that illion has NO equivalent for.
# Rows with these labels are excluded from agreement-rate calculation.
UNMAPPABLE_OUR_CATS: frozenset[str] = frozenset({
    "Debt Consolidation",
})

# ── reverse mapping (illion → our labels) for human review ─────────
# 1:1 exact-name matches are marked as "[同名]" for readability.

ILLION_TO_OUR: dict[str, list[str]] = {
    # -- exact name match (1:1) --
    "Internal Transfer":            ["[同名] Internal Transfer"],
    "External Transfers":           ["[同名] External Transfers"],
    "Dining Out":                   ["[同名] Dining Out"],
    "Retail":                       ["[同名] Retail"],
    "Groceries":                    ["[同名] Groceries"],
    "Health":                       ["[同名] Health"],
    "Automotive":                   ["[同名] Automotive"],
    "Entertainment":                ["[同名] Entertainment"],
    "Home Improvement":             ["[同名] Home Improvement"],
    "Travel":                       ["[同名] Travel"],
    "Information":                  ["[同名] Information"],
    "Personal Care":                ["[同名] Personal Care"],
    "Transport":                    ["[同名] Transport"],
    "Education":                    ["[同名] Education"],
    "Gambling":                     ["[同名] Gambling"],
    "Gyms and other memberships":   ["[同名] Gyms and other memberships"],
    "Pet Care":                     ["[同名] Pet Care"],
    "Donations":                    ["[同名] Donations"],
    "Utilities":                    ["[同名] Utilities"],
    "Telecommunications":           ["[同名] Telecommunications"],
    "Rent":                         ["[同名] Rent"],
    "Department Stores":            ["[同名] Department Stores"],
    "Insurance":                    ["[同名] Insurance"],
    "Subscription TV":              ["[同名] Subscription TV"],
    "Dishonours":                   ["[同名] Dishonours"],
    "Debt Collection":              ["[同名] Debt Collection"],
    "Overdrawn":                    ["[同名] Overdrawn"],
    "SACC Loans":                   ["[同名] SACC Loans"],

    # -- 1:1, different names --
    "Fees":                         ["fee"],
    "Centrelink":                   ["centrelink"],
    "Credit Card Repayments":       ["Credit Card Repayment"],

    # -- 1:N: illion lumps together our fine-grained categories --
    "Wages":                        ["salary_payg", "salary_packaging", "self_employed_gig"],
    "Non SACC Loans":               ["BNPL", "Wage Advance", "Non SACC Loans",
                                     "Personal Loan Unknown", "Contract Loans", "LOC",
                                     "Home Loan", "Car Loan"],

    # -- illion categories with no direct equivalent in our engine --
    "All Other Credits":            [],   # 我们没有兜底收入类，可能分散在多个标签
}

# Columns to hide in the detail/confusion sheets.
_HIDDEN_COLS = frozenset({
    "sample_datetime", "job_id", "transaction_id", "account_type",
    "credit_limit", "trx_type", "bsb", "account_no",
    "illion_trx_uuid", "balance",
})


# ── data loading ─────────────────────────────────────────────────────────

def _load(report_path: Path) -> pd.DataFrame:
    our = pd.read_excel(report_path, sheet_name="transactions")
    illion = pd.read_csv(SAMPLE_CSV, encoding="utf-8-sig")
    df = our.copy()
    df["_illion_cat"] = illion["category"].values
    df["_illion_cp"] = illion["third_party"].values
    return df


# ── helpers ──────────────────────────────────────────────────────────────

def _map(our_cat: str) -> str:
    c = str(our_cat).strip()
    return OUR_TO_ILLION.get(c, c)


def _agree(our: str, illion_cat: str) -> bool | None:
    """True when our mapped category matches illion's category exactly.

    Returns None when our category has no illion equivalent (unmappable).
    """
    our, illion_cat = str(our).strip(), str(illion_cat).strip()
    if our in UNMAPPABLE_OUR_CATS:
        return None
    return _map(our) == illion_cat


def _is_classified(series: pd.Series) -> pd.Series:
    return (
        series.notna()
        & series.astype(str).str.strip().ne("")
        & series.astype(str).str.strip().ne("unclassified")
    )


# ── excel helpers ────────────────────────────────────────────────────────

def _auto_width(worksheet, min_width: int = 8, max_width: int = 40) -> None:
    """Set column widths based on content."""
    for col_cells in worksheet.columns:
        col_letter = col_cells[0].column_letter
        max_chars = 0
        for cell in col_cells:
            if cell.value is not None:
                max_chars = max(max_chars, len(str(cell.value)))
        width = max(min_width, min(max_chars + 2, max_width))
        worksheet.column_dimensions[col_letter].width = width


def _hide_columns(worksheet) -> None:
    """Auto-hide noisy detail columns on the transactions sheet."""
    col_map: dict[str, int] = {}
    for cell in worksheet[1]:
        if cell.value is not None:
            col_map[str(cell.value)] = cell.column

    for col_name in _HIDDEN_COLS:
        col_idx = col_map.get(col_name)
        if col_idx is not None:
            worksheet.column_dimensions[
                worksheet.cell(row=1, column=col_idx).column_letter
            ].hidden = True


# ── colour helpers ──────────────────────────────────────────────────────

_AGREE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_DISAGREE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_WARN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_OUR_COL_FILL = PatternFill(start_color="F4B4C2", end_color="F4B4C2", fill_type="solid")
_IL_COL_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_header_row(worksheet) -> None:
    """Apply dark-blue header with white bold text to row 1."""
    for cell in worksheet[1]:
        if cell.value is not None:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT


def _colour_confusion_pivot(worksheet, meta_cols: list[str]) -> None:
    """Colour the confusion pivot: green for agreement cells, red for off-diagonal.

    For each row (our category), the mapped illion category column receives
    green; all other illion-category columns receive light red.
    """
    # Build a col-name → col-index map from the header row.
    header: dict[str, int] = {}
    for cell in worksheet[1]:
        if cell.value is not None:
            header[str(cell.value)] = cell.column

    meta_col_names = set(meta_cols)
    cat_cols = [
        (col_name, col_idx)
        for col_name, col_idx in header.items()
        if col_name not in meta_col_names
    ]

    for row in range(2, worksheet.max_row + 1):
        our_cat = str(worksheet.cell(row=row, column=1).value or "")
        mapped = OUR_TO_ILLION.get(our_cat, our_cat)  # our → illion name

        for col_name, col_idx in cat_cols:
            cell = worksheet.cell(row=row, column=col_idx)
            val = cell.value
            if val is None:
                continue
            try:
                n = int(val)
            except (ValueError, TypeError):
                continue
            if n == 0:
                continue
            if col_name == mapped:
                cell.fill = _AGREE_FILL
            else:
                cell.fill = _DISAGREE_FILL


def _colour_category_summary(worksheet) -> None:
    """Colour the agreement-rate column: green ≥80%, yellow 50-79%, red <50%."""
    # Find the 一致率 column index.
    rate_col: int | None = None
    for cell in worksheet[1]:
        if cell.value and "一致率" in str(cell.value):
            rate_col = cell.column
            break
    if rate_col is None:
        return

    for row in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row, column=rate_col)
        val = str(cell.value or "")
        if val in ("—", "") or "N/A" in val:
            continue
        try:
            pct = float(val.rstrip("%"))
        except ValueError:
            continue
        if pct >= 80:
            cell.fill = _AGREE_FILL
        elif pct >= 50:
            cell.fill = _WARN_FILL
        else:
            cell.fill = _DISAGREE_FILL


def _colour_disagreement_flow(worksheet) -> None:
    """Colour flow rows: green when our mapped category matches illion, red otherwise.

    Rows where our category has no illion equivalent (unmappable) are left uncoloured.
    """
    our_col: int | None = None
    il_col: int | None = None
    for cell in worksheet[1]:
        if cell.value == "我们的分类":
            our_col = cell.column
        elif cell.value == "illion分类":
            il_col = cell.column

    if our_col is None or il_col is None:
        return

    for row in range(2, worksheet.max_row + 1):
        our_cat = str(worksheet.cell(row=row, column=our_col).value or "")
        il_cat = str(worksheet.cell(row=row, column=il_col).value or "")
        if our_cat in UNMAPPABLE_OUR_CATS:
            # unmappable — leave neutral (no colour)
            continue
        mapped = OUR_TO_ILLION.get(our_cat, our_cat)
        if mapped == il_cat:
            for col in range(1, worksheet.max_column + 1):
                worksheet.cell(row=row, column=col).fill = _AGREE_FILL
        else:
            for col in range(1, worksheet.max_column + 1):
                worksheet.cell(row=row, column=col).fill = _DISAGREE_FILL


def _colour_disagreement_detail(worksheet) -> None:
    """Colour our-category and illion-category columns differently."""
    our_col: int | None = None
    il_col: int | None = None
    for cell in worksheet[1]:
        if cell.value == "finv_category":
            our_col = cell.column
        elif cell.value and "illion_category" in str(cell.value):
            il_col = cell.column

    if our_col is None and il_col is None:
        return

    for row in range(2, worksheet.max_row + 1):
        if our_col is not None:
            worksheet.cell(row=row, column=our_col).fill = _OUR_COL_FILL
        if il_col is not None:
            worksheet.cell(row=row, column=il_col).fill = _IL_COL_FILL


def _write_sheet(writer, name: str, df: pd.DataFrame) -> None:
    """Write a DataFrame as a sheet and apply basic formatting."""
    df.to_excel(writer, sheet_name=name, index=True)
    ws = writer.book[name]
    ws.freeze_panes = "B2" if df.index.name else "A2"
    ws.auto_filter.ref = ws.dimensions
    _auto_width(ws)
    if name == "disagreement_detail":
        _hide_columns(ws)


# ── main entry point ─────────────────────────────────────────────────────

def run(
    report_path: str | Path = DEFAULT_OUTPUT,
    result_path: str | Path = DEFAULT_RESULT,
) -> Path:
    report_path = Path(report_path)
    result_path = Path(result_path)
    result_path.parent.mkdir(parents=True, exist_ok=True)

    df = _load(report_path)

    our_fc = df["finv_category"]
    illion_cat = df["_illion_cat"]
    our_cp = df["counterparty"]
    illion_cp = df["_illion_cp"]

    our_fc_ok = _is_classified(our_fc)
    illion_ok = _is_classified(illion_cat)
    our_cp_ok = _is_classified(our_cp)
    illion_cp_ok = _is_classified(illion_cp)

    n = len(df)
    both_mask = our_fc_ok & illion_ok

    # ── engine ownership: which engine produced each category ────────────
    _cat_engine: dict[str, str] = {}
    if "classification_engine" in df.columns:
        for cat_name in our_fc[our_fc_ok].unique():
            engines = df.loc[our_fc_ok & (our_fc == cat_name), "classification_engine"]
            primary = engines.value_counts().index[0] if len(engines) > 0 else "?"
            _cat_engine[cat_name] = primary

    # ── 1. Coverage stats ─────────────────────────────────────────────────
    coverage_rows = [
        {
            "指标": "finv_category / category",
            "我们已分类": f"{our_fc_ok.sum()} ({our_fc_ok.sum()/n*100:.1f}%)",
            "illion已分类": f"{illion_ok.sum()} ({illion_ok.sum()/n*100:.1f}%)",
            "总行数": n,
        },
        {
            "指标": "counterparty / third_party",
            "我们已分类": f"{our_cp_ok.sum()} ({our_cp_ok.sum()/n*100:.1f}%)",
            "illion已分类": f"{illion_cp_ok.sum()} ({illion_cp_ok.sum()/n*100:.1f}%)",
            "总行数": n,
        },
        {
            "指标": "双方都未分类",
            "我们已分类": "",
            "illion已分类": "",
            "总行数": int((~our_fc_ok & ~illion_ok).sum()),
        },
    ]
    coverage_df = pd.DataFrame(coverage_rows)

    # ── 2. Agreement summary ──────────────────────────────────────────────
    n_both = int(both_mask.sum())
    # Exclude rows where our category has no illion equivalent.
    comparable_mask = both_mask & ~our_fc.isin(UNMAPPABLE_OUR_CATS)
    n_comparable = int(comparable_mask.sum())
    n_unmappable = n_both - n_comparable

    agree_results = [
        _agree(our_fc[i], illion_cat[i])
        for i in df.index[comparable_mask]
    ]
    agree_count = sum(1 for v in agree_results if v is True)
    disagree_count = sum(1 for v in agree_results if v is False)

    agreement_rows = [
        {"指标": "双方都有标签的行", "值": f"{n_both:,}"},
        {"指标": "  其中可比较 (illion有对应类别)", "值": f"{n_comparable:,}"},
        {"指标": "  其中不可比较 (illion无对应类别)", "值": f"{n_unmappable:,}"},
        {"指标": "一致 (映射后严格匹配)", "值": f"{agree_count:,} / {n_comparable:,} = {agree_count/n_comparable*100:.1f}%" if n_comparable > 0 else "—"},
        {"指标": "不一致", "值": f"{disagree_count:,} / {n_comparable:,} = {disagree_count/n_comparable*100:.1f}%" if n_comparable > 0 else "—"},
        {"指标": "一致 (vs 我们已分类总数)", "值": f"{agree_count:,} / {int(our_fc_ok.sum()):,} = {agree_count/our_fc_ok.sum()*100:.1f}%"},
    ]
    agreement_df = pd.DataFrame(agreement_rows)

    # ── 3. Confusion pivot ────────────────────────────────────────────────
    pivot_rows = []
    for our_cat_name in sorted(our_fc[our_fc_ok].unique()):
        mask = our_fc_ok & (our_fc == our_cat_name)
        total_n = int(mask.sum())
        both_for_cat = mask & illion_ok
        illion_unclassified_n = int((mask & ~illion_ok).sum())

        row: dict = {
            "finv_category": our_cat_name,
            "total": total_n,
            "illion_unclassified": illion_unclassified_n,
        }

        if total_n > 0:
            if our_cat_name in UNMAPPABLE_OUR_CATS:
                row["agree_pct"] = None  # 不可比较：illion 无对应类别
            else:
                ag = sum(
                    1 for i in df.index[both_for_cat]
                    if _agree(our_cat_name, illion_cat[i]) is True
                )
                # 分母 = total（我们分的全部），illion 未覆盖的视为不一致
                row["agree_pct"] = round(ag / total_n * 100, 1)
        else:
            row["agree_pct"] = None

        for illion_val, cnt in (
            df.loc[mask & illion_ok, "_illion_cat"].value_counts().items()
        ):
            row[str(illion_val)] = cnt
        pivot_rows.append(row)
    pivot = (
        pd.DataFrame(pivot_rows)
        .set_index("finv_category")
        .fillna(0)
    )
    meta_cols = ["total", "illion_unclassified", "agree_pct"]
    cat_cols = sorted(
        [c for c in pivot.columns if c not in meta_cols],
        key=lambda x: pivot[x].sum(),
        reverse=True,
    )
    pivot = pivot[meta_cols + cat_cols]
    # Convert float counts to int.
    for c in cat_cols:
        pivot[c] = pivot[c].astype(int)
    pivot["total"] = pivot["total"].astype(int)
    pivot["illion_unclassified"] = pivot["illion_unclassified"].astype(int)

    # ── 4. Counterparty comparison ────────────────────────────────────────
    cp_both = our_cp_ok & illion_cp_ok
    n_cp_both = int(cp_both.sum())
    cp_agree = (
        our_cp[cp_both].astype(str).str.strip()
        == illion_cp[cp_both].astype(str).str.strip()
    ).sum()
    counterparty_rows = [
        {"指标": "双方都有counterparty", "值": f"{n_cp_both:,}"},
        {"指标": "完全一致", "值": f"{cp_agree:,} / {n_cp_both:,} = {cp_agree/n_cp_both*100:.1f}%"},
        {"指标": "我们独有", "值": int(our_cp_ok.sum() - n_cp_both)},
        {"指标": "illion独有", "值": int(illion_cp_ok.sum() - n_cp_both)},
    ]
    counterparty_df = pd.DataFrame(counterparty_rows)

    # ── 5. Disagreement flow ───────────────────────────────────────────────
    # For each of our categories, show what illion calls the same transactions.
    # Long-form table — easier to scan than the dense confusion pivot.
    flow_rows: list[dict] = []
    for our_cat_name in sorted(our_fc[our_fc_ok].unique()):
        mask = our_fc_ok & (our_fc == our_cat_name) & illion_ok
        total_n = int(mask.sum())
        if total_n == 0:
            continue
        total_amt = df.loc[mask, "amount"].abs().sum()
        for il_cat, cnt in df.loc[mask, "_illion_cat"].value_counts().items():
            sub_mask = mask & (illion_cat == il_cat)
            amt = df.loc[sub_mask, "amount"].abs().sum()
            flow_rows.append({
                "我们的分类": our_cat_name,
                "illion分类": str(il_cat),
                "笔数": cnt,
                "行占比": round(cnt / total_n * 100, 1),
                "金额合计": round(amt, 2),
                "金额占比": round(amt / total_amt * 100, 1) if total_amt > 0 else 0.0,
            })
    flow_df = pd.DataFrame(flow_rows)
    flow_df = flow_df.sort_values(["我们的分类", "笔数"], ascending=[True, False])

    # ── 6. Coverage gaps ──────────────────────────────────────────────────
    our_only = our_fc[our_fc_ok & ~illion_ok].value_counts().reset_index()
    our_only.columns = ["finv_category", "count"]
    illion_only = (
        illion_cat[~our_fc_ok & illion_ok].value_counts().reset_index()
    )
    illion_only.columns = ["illion_category", "count"]

    # ── 7. Per-category agreement breakdown ───────────────────────────────
    cat_summary_rows = []
    for our_cat_name in pivot.index:
        r = pivot.loc[our_cat_name]
        if our_cat_name in UNMAPPABLE_OUR_CATS:
            agree_str = "N/A (illion无此类别)"
        elif r["agree_pct"] and r["agree_pct"] > 0:
            agree_str = f"{r['agree_pct']:.0f}%"
        else:
            agree_str = "—"
        total_n = int(r["total"])
        unclassified_n = int(r.get("illion_unclassified", 0))
        illion_top = []
        for c in cat_cols:
            v = int(r[c])
            if v > 0:
                illion_top.append(f"{c}({v})")
                if len(illion_top) >= 3:
                    break
        cat_summary_rows.append({
            "我们的分类": our_cat_name,
            "引擎": _cat_engine.get(our_cat_name, "—"),
            "数量": total_n,
            "illion未覆盖": unclassified_n,
            "一致率": agree_str,
            "illion对应标签 (top 3)": ", ".join(illion_top) if illion_top else "—",
        })
    cat_summary_df = pd.DataFrame(cat_summary_rows)

    # ── 8. Disagreement detail ────────────────────────────────────────────
    agree_series = pd.Series(
        [
            _agree(our_fc[i], illion_cat[i])
            for i in range(len(df))
        ],
        index=df.index,
    )
    # exclude unmappable rows (None) — they are not disagreements
    disagree_mask = both_mask & agree_series.eq(False)
    disagree_cols = [
        "user_id", "application_id", "transaction_date", "amount",
        "dr_cr", "text",
        "finv_category", "_illion_cat",
        "counterparty", "_illion_cp",
        "classification_rule_id",
    ]
    available = [c for c in disagree_cols if c in df.columns]
    disagree_df = df.loc[disagree_mask, available].copy()
    disagree_df.rename(columns={
        "_illion_cat": "illion_category",
        "_illion_cp": "illion_counterparty",
    }, inplace=True)
    disagree_df = disagree_df.sort_values(
        ["finv_category", "illion_category"]
    )

    # ═══════════════════════════════════════════════════════════════════
    #  WRITE XLSX
    # ═══════════════════════════════════════════════════════════════════
    with pd.ExcelWriter(result_path, engine="openpyxl") as writer:
        _write_sheet(writer, "coverage", coverage_df.set_index("指标"))
        _write_sheet(writer, "agreement", agreement_df.set_index("指标"))
        _write_sheet(writer, "confusion", pivot)
        _colour_confusion_pivot(writer.book["confusion"], meta_cols)
        _write_sheet(writer, "category_summary", cat_summary_df.set_index("我们的分类"))
        _colour_category_summary(writer.book["category_summary"])
        _write_sheet(writer, "counterparty", counterparty_df.set_index("指标"))
        if len(flow_df) > 0:
            _write_sheet(writer, "disagreement_flow", flow_df.set_index("我们的分类"))
            _colour_disagreement_flow(writer.book["disagreement_flow"])
        if len(our_only) > 0:
            _write_sheet(writer, "our_only", our_only.set_index("finv_category"))
        if len(illion_only) > 0:
            _write_sheet(writer, "illion_only", illion_only.set_index("illion_category"))
        if len(disagree_df) > 0:
            _write_sheet(writer, "disagreement_detail", disagree_df)
            _colour_disagreement_detail(writer.book["disagreement_detail"])
        # reverse mapping sheet for human review
        illion_to_our_rows = [
            {"illion_category": k, "our_labels": ", ".join(v) if v else "（无对应）"}
            for k, v in ILLION_TO_OUR.items()
        ]
        _write_sheet(writer, "illion_to_our_mapping",
                     pd.DataFrame(illion_to_our_rows).set_index("illion_category"))
        # Style headers for all sheets
        for name in writer.book.sheetnames:
            _style_header_row(writer.book[name])

    print(f"→ 对比报告已写入: {result_path}")
    print(f"  覆盖: 我们 {our_fc_ok.sum()/n*100:.1f}% | illion {illion_ok.sum()/n*100:.1f}%")
    print(f"  一致率: {agree_count/n_comparable*100:.1f}% ({agree_count:,}/{n_comparable:,})"
          f"  [可比较行, 排除unmappable]")
    print(f"  不一致: {disagree_count:,} 行")
    return result_path


# ── CLI ──────────────────────────────────────────────────────────────────

def main() -> None:
    p = argparse.ArgumentParser(
        description="Compare our labels against illion labels"
    )
    p.add_argument(
        "--output",
        default=str(DEFAULT_OUTPUT),
        help="Path to classification_report.xlsx (our output)",
    )
    p.add_argument(
        "--result",
        default=str(DEFAULT_RESULT),
        help="Path to write compare_result.xlsx",
    )
    args = p.parse_args()
    run(args.output, args.result)


if __name__ == "__main__":
    main()
