from __future__ import annotations

from collections.abc import Iterable

from openpyxl.styles import Alignment, Font, PatternFill


DEFAULT_FONT_NAME = "Microsoft YaHei"
HEADER_FILL = "FF000000"
HEADER_FONT_COLOR = "FFFFFFFF"

# Header colours for highlighted columns.
RED_FILL = PatternFill(fill_type="solid", fgColor="FFE0B4B4")
GREEN_FILL = PatternFill(fill_type="solid", fgColor="FFB4E0B4")


def _apply_base_format(worksheet) -> None:
    """Apply freeze panes, auto-filter, and header styling.

    Body cells use Excel defaults (no per-cell formatting) to avoid
    O(rows × cols) overhead on large sheets.
    """
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    header_font = Font(
        name=DEFAULT_FONT_NAME,
        color=HEADER_FONT_COLOR,
        bold=True,
    )
    header_fill = PatternFill(fill_type="solid", fgColor=HEADER_FILL)

    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def format_sheets(workbook, sheet_names: Iterable[str]) -> None:
    for sheet_name in sheet_names:
        _apply_base_format(workbook[sheet_name])


def format_transactions_sheet(
    worksheet,
    *,
    hidden_cols: frozenset[str] | None = None,
    red_cols: frozenset[str] | None = None,
    green_cols: frozenset[str] | None = None,
) -> None:
    """Format the transactions sheet with base styling, hidden columns, and
    coloured header highlights.

    Hidden columns are still present in the file — the user can manually
    unhide them in Excel.
    """
    _apply_base_format(worksheet)

    if hidden_cols is None:
        hidden_cols = frozenset()
    if red_cols is None:
        red_cols = frozenset()
    if green_cols is None:
        green_cols = frozenset()

    # ── build column-name → column-index map from the header row ──
    col_map: dict[str, int] = {}
    for cell in worksheet[1]:
        if cell.value is not None:
            col_map[str(cell.value)] = cell.column

    # ── hide columns ──
    for col_name in hidden_cols:
        col_idx = col_map.get(col_name)
        if col_idx is not None:
            worksheet.column_dimensions[
                worksheet.cell(row=1, column=col_idx).column_letter
            ].hidden = True

    # ── colour header cells ──
    for col_name in red_cols:
        col_idx = col_map.get(col_name)
        if col_idx is not None:
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = RED_FILL
            cell.font = Font(
                name=DEFAULT_FONT_NAME,
                color="FFAA0000",
                bold=True,
            )

    for col_name in green_cols:
        col_idx = col_map.get(col_name)
        if col_idx is not None:
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = GREEN_FILL
            cell.font = Font(
                name=DEFAULT_FONT_NAME,
                color="FF006600",
                bold=True,
            )
